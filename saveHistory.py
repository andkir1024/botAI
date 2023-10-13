import os
from datetime import datetime
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from managerQR import managerQR
from okDeskUtils import okDesk
from processorMenu import *
from aiogram.types import InputFile

import pandas as pd
from openpyxl import Workbook, load_workbook

class history:
    def __init__(self):
        pass
    def save(userInfo, msg: types.Message):
        supportMode = userInfo.supportMode.lower()
        info = msg.text
        dir_path = os.path.dirname(os.path.realpath(__file__))
        dirHistory = dir_path + mainConst.DIR_HISTORY
        if not os.path.isdir(dirHistory):
            os.mkdir(dirHistory)

        now = datetime.now() 
        current_time = now.strftime("%Y:%d:%m:%H:%M:%S")
        current_file_name = now.strftime("%Y_%d_%m")
        
        xlName = dirHistory + current_file_name + '.xlsx'
        # xlName = dirHistory + 'example.xlsx'
        try:        
            wb = load_workbook(xlName)
            sheet = wb.active
            worker = userInfo.first_name + ' ' + userInfo.last_name
            request = "Проблема с поддержкой"
            if supportMode == 'menuLogistic'.lower():
                request = "Проблема с логистикой"
            if supportMode == 'menuWareHouse'.lower():
                request = "Проблема со складом"
            if supportMode == 'menuSoftError'.lower():
                request = "Сообщить об ошибке ПО"
            new_row = [current_time, request, worker, info]
            sheet.append(new_row)
            wb.save(xlName)
        except:
            my_wb = Workbook()
            my_sheet = my_wb.active
            new_row = ['Дата', 'Запрос', 'Сотрудник', 'Сообщение']
            my_sheet.append(new_row)
            # my_sheet.column_dimensions['A1'].width = 200
            my_sheet.append(history.creatRow(current_time, supportMode, userInfo, info))
            my_wb.save(xlName)
            pass
        return
        
    def creatRow(current_time, supportMode, userInfo, info):
        worker = userInfo.first_name + ' ' + userInfo.last_name
        request = "Проблема с поддержкой"
        if supportMode == 'menuLogistic'.lower():
            request = "Проблема с логистикой"
        if supportMode == 'menuWareHouse'.lower():
            request = "Проблема со складом"
        if supportMode == 'menuSoftError'.lower():
            request = "Сообщить об ошибке ПО"
        new_row = [current_time, request, worker, info]
        return new_row
